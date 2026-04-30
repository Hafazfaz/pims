from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, View
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import AuditLogEntry
from user_management.models import CustomUser

class AuditLogListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = AuditLogEntry
    template_name = 'audit_log/audit_log_list.html'
    context_object_name = 'log_entries'
    paginate_by = 20

    def test_func(self):
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return True
        return user.groups.filter(name__iexact="Executives").exists()

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'user', 'user__staff__department', 'user__staff__designation'
        )

        # Filtering
        user_id = self.request.GET.get('user')
        action = self.request.GET.get('action')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        search_query = self.request.GET.get('q')

        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if action:
            queryset = queryset.filter(action=action)
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        if end_date:
            queryset = queryset.filter(timestamp__lte=end_date)
        if search_query:
            queryset = queryset.filter(
                Q(details__icontains=search_query) |
                Q(action__icontains=search_query) |
                Q(user__username__icontains=search_query) |
                Q(ip_address__icontains=search_query) |
                Q(user_agent__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['all_users'] = CustomUser.objects.all().order_by('username')
        context['all_actions'] = sorted(list(set(AuditLogEntry.objects.values_list('action', flat=True))))
        
        user_id = self.request.GET.get('user', '')
        context['selected_user'] = int(user_id) if user_id.isdigit() else ''
        
        context['selected_action'] = self.request.GET.get('action', '')
        context['selected_start_date'] = self.request.GET.get('start_date', '')
        context['selected_end_date'] = self.request.GET.get('end_date', '')
        context['search_query'] = self.request.GET.get('q', '')
        
        # Preserve query parameters for pagination
        params = self.request.GET.copy()
        if 'page' in params:
            del params['page']
        context['query_params'] = '&' + params.urlencode() if params else ''
        
        return context


class MyActivityReportView(LoginRequiredMixin, View):
    """Personal activity report — aggregates the user's own audit log into meaningful sections."""

    # Action groups
    APPROVAL_ACTIONS = {'DOCUMENT_APPROVED', 'CHAIN_STEP_APPROVED'}
    REJECTION_ACTIONS = {'DOCUMENT_REJECTED', 'CHAIN_STEP_REJECTED'}
    ACCESS_GRANTED_ACTIONS = {'ACCESS_REQUEST_APPROVED'}
    ACCESS_DENIED_ACTIONS = {'ACCESS_REQUEST_REJECTED'}
    FILE_MOVEMENT_ACTIONS = {'FILE_SENT', 'DOCUMENT_FORWARDED', 'MOVEMENT_CLOSED'}
    FILE_ACTIONS = {'FILE_CREATED', 'FILE_UPDATED', 'FILE_CLOSED', 'FILE_ARCHIVED', 'FILE_ACTIVATED'}

    def get(self, request):
        user = request.user
        days = int(request.GET.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        qs = AuditLogEntry.objects.filter(user=user, timestamp__gte=since).order_by('-timestamp')

        def group(actions):
            return qs.filter(action__in=actions)

        # Summary counts
        summary = {
            'approvals': group(self.APPROVAL_ACTIONS).count(),
            'rejections': group(self.REJECTION_ACTIONS).count(),
            'access_granted': group(self.ACCESS_GRANTED_ACTIONS).count(),
            'access_denied': group(self.ACCESS_DENIED_ACTIONS).count(),
            'files_moved': group(self.FILE_MOVEMENT_ACTIONS).count(),
            'file_actions': group(self.FILE_ACTIONS).count(),
            'total': qs.count(),
        }

        # Action breakdown for chart
        breakdown = (
            qs.values('action')
            .annotate(count=Count('action'))
            .order_by('-count')[:10]
        )

        return render(request, 'audit_log/my_activity_report.html', {
            'summary': summary,
            'breakdown': breakdown,
            'approvals': group(self.APPROVAL_ACTIONS).select_related('user')[:20],
            'rejections': group(self.REJECTION_ACTIONS).select_related('user')[:20],
            'access_granted': group(self.ACCESS_GRANTED_ACTIONS).select_related('user')[:20],
            'access_denied': group(self.ACCESS_DENIED_ACTIONS).select_related('user')[:20],
            'file_movements': group(self.FILE_MOVEMENT_ACTIONS).select_related('user')[:20],
            'days': days,
        })


class MyActivityReportView(LoginRequiredMixin, View):
    """
    Personal activity report for the requesting user.
    Registry can also view any user's report via ?user_id=<pk>.
    """

    APPROVAL_ACTIONS = {'DOCUMENT_APPROVED', 'CHAIN_STEP_APPROVED'}
    REJECTION_ACTIONS = {'DOCUMENT_REJECTED', 'CHAIN_STEP_REJECTED'}
    ACCESS_GRANTED_ACTIONS = {'ACCESS_REQUEST_APPROVED'}
    ACCESS_DENIED_ACTIONS = {'ACCESS_REQUEST_REJECTED'}
    FILE_MOVEMENT_ACTIONS = {'FILE_SENT', 'DOCUMENT_FORWARDED', 'MOVEMENT_CLOSED'}
    FILE_ACTIONS = {'FILE_CREATED', 'FILE_UPDATED', 'FILE_CLOSED', 'FILE_ARCHIVED', 'FILE_ACTIVATED'}

    def get(self, request):
        staff = getattr(request.user, 'staff', None)
        is_registry = staff and staff.is_registry

        # Registry: default to all users; can filter by searched user
        target_user = None if is_registry else request.user
        viewed_user = None
        search_q = request.GET.get('q', '').strip()
        user_id = request.GET.get('user_id')

        if user_id:
            try:
                target_user = CustomUser.objects.get(pk=user_id)
                viewed_user = target_user
            except CustomUser.DoesNotExist:
                pass
        elif search_q and is_registry:
            from django.db.models import Q as DQ
            match = CustomUser.objects.filter(
                DQ(first_name__icontains=search_q) |
                DQ(last_name__icontains=search_q) |
                DQ(email__icontains=search_q) |
                DQ(username__icontains=search_q)
            ).first()
            if match:
                target_user = match
                viewed_user = match
        elif not is_registry:
            target_user = request.user

        days = request.GET.get('days')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        qs = AuditLogEntry.objects.all().order_by('-timestamp')

        if start_date:
            qs = qs.filter(timestamp__date__gte=start_date)
        if end_date:
            qs = qs.filter(timestamp__date__lte=end_date)
        
        # Fallback to days if no specific range is provided
        if not start_date and not end_date:
            days_int = int(days) if days and days.isdigit() else 30
            since = timezone.now() - timedelta(days=days_int)
            qs = qs.filter(timestamp__gte=since)
        else:
            days_int = None # Mark that we are using a custom range
        if target_user:
            qs = qs.filter(user=target_user)

        def group(actions):
            return qs.filter(action__in=actions)

        summary = {
            'approvals': group(self.APPROVAL_ACTIONS).count(),
            'rejections': group(self.REJECTION_ACTIONS).count(),
            'access_granted': group(self.ACCESS_GRANTED_ACTIONS).count(),
            'access_denied': group(self.ACCESS_DENIED_ACTIONS).count(),
            'files_moved': group(self.FILE_MOVEMENT_ACTIONS).count(),
            'file_actions': group(self.FILE_ACTIONS).count(),
            'total': qs.count(),
        }

        # Action breakdown with top users per action
        breakdown_qs = (
            qs.values('action')
            .annotate(count=Count('action'))
            .order_by('-count')[:10]
        )
        
        breakdown = []
        for item in breakdown_qs:
            top_users = (
                qs.filter(action=item['action'])
                .values('user__username', 'user__first_name', 'user__last_name')
                .annotate(user_count=Count('user'))
                .order_by('-user_count')[:3]
            )
            users_list = []
            for u in top_users:
                name = f"{u['user__first_name']} {u['user__last_name']}".strip() or u['user__username'] or "System"
                users_list.append(f"{name} ({u['user_count']})")
            
            item['top_users'] = ", ".join(users_list)
            breakdown.append(item)

        all_users = CustomUser.objects.all().order_by('username') if is_registry else None

        viewed_user_staff = None
        if viewed_user:
            try:
                viewed_user_staff = viewed_user.staff
            except Exception:
                pass

        return render(request, 'audit_log/my_activity_report_partial.html' if request.headers.get('HX-Request') else 'audit_log/my_activity_report.html', {
            'summary': summary,
            'breakdown': breakdown,
            'approvals': group(self.APPROVAL_ACTIONS)[:50],
            'rejections': group(self.REJECTION_ACTIONS)[:50],
            'access_granted': group(self.ACCESS_GRANTED_ACTIONS)[:50],
            'access_denied': group(self.ACCESS_DENIED_ACTIONS)[:50],
            'file_movements': group(self.FILE_MOVEMENT_ACTIONS)[:50],
            'file_actions': group(self.FILE_ACTIONS)[:50],
            'days': days_int,
            'start_date': start_date,
            'end_date': end_date,
            'is_registry': is_registry,
            'all_users': all_users,
            'viewed_user': viewed_user,
            'viewed_user_staff': viewed_user_staff,
            'search_q': search_q,
        })


class ActivityUserSearchView(LoginRequiredMixin, View):
    """HTMX: returns a dropdown of matching users for registry search."""

    def get(self, request):
        staff = getattr(request.user, 'staff', None)
        if not (staff and staff.is_registry):
            return HttpResponse('')
        q = request.GET.get('q', '').strip()
        if len(q) < 2:
            return HttpResponse('')
        from django.db.models import Q as DQ
        users = CustomUser.objects.filter(
            DQ(first_name__icontains=q) | DQ(last_name__icontains=q) |
            DQ(email__icontains=q) | DQ(username__icontains=q)
        ).select_related('staff__designation', 'staff__department')[:8]

        html = '<div class="absolute z-50 w-full mt-1 bg-white border border-slate-200 rounded-lg shadow-lg overflow-hidden">'
        for u in users:
            name = u.get_full_name() or u.username
            s = getattr(u, 'staff', None)
            sub = ' · '.join(filter(None, [
                s.designation.name if s and s.designation else '',
                s.department.name if s and s.department else '',
            ]))
            safe_name = name.replace("'", "\\'")
            html += f'''<div class="px-4 py-3 hover:bg-slate-50 cursor-pointer border-b border-slate-100 last:border-0"
                             hx-get="/audit/my-activity/?user_id={u.pk}&days={request.GET.get('days', 30)}"
                             hx-target="#activity-content"
                             hx-push-url="true"
                             hx-indicator="#activity-content"
                             @click="query='{safe_name}'; loading=true; showResults=false">
                <p class="text-xs font-bold text-slate-900">{name}</p>
                <p class="text-[10px] text-slate-400">{u.email}{f" · {sub}" if sub else ""}</p>
            </div>'''
        html += '</div>'
        return HttpResponse(html)


class AccessDeniedExportView(LoginRequiredMixin, View):
    """Export access denied log entries as CSV or Excel for the current/viewed user."""

    def get(self, request):
        import csv
        from django.http import HttpResponse as HR
        from datetime import timedelta

        staff = getattr(request.user, 'staff', None)
        is_registry = staff and staff.is_registry

        target_user = request.user
        user_id = request.GET.get('user_id')
        if user_id and is_registry:
            try:
                target_user = CustomUser.objects.get(pk=user_id)
            except CustomUser.DoesNotExist:
                pass

        days = int(request.GET.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        entries = AuditLogEntry.objects.filter(
            action='ACCESS_REQUEST_REJECTED',
            timestamp__gte=since,
        ).order_by('-timestamp')

        if not is_registry:
            entries = entries.filter(user=target_user)

        fmt = request.GET.get('format', 'csv')

        rows = []
        for e in entries:
            d = e.details or {}
            rows.append({
                'Date': e.timestamp.strftime('%Y-%m-%d %H:%M'),
                'Denied By': d.get('denied_by', e.user.get_full_name() if e.user else ''),
                'Requested By': d.get('requested_by', ''),
                'Staff ID': d.get('staff_id', ''),
                'File Number': d.get('file', ''),
                'File Title': d.get('file_title', ''),
                'Reason': d.get('denial_reason', ''),
            })

        if fmt == 'excel':
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Access Denied'
                if rows:
                    ws.append(list(rows[0].keys()))
                    for r in rows:
                        ws.append(list(r.values()))
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                response = HR(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="access_denied_{timezone.now().date()}.xlsx"'
                return response
            except ImportError:
                pass  # fall through to CSV

        response = HR(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="access_denied_{timezone.now().date()}.csv"'
        if rows:
            writer = csv.DictWriter(response, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        return response


class FullActivityExportView(LoginRequiredMixin, View):
    """Export activity log entries with comprehensive details as CSV or Excel."""

    SECTION_MAP = {
        'approvals': {'DOCUMENT_APPROVED', 'CHAIN_STEP_APPROVED'},
        'rejections': {'DOCUMENT_REJECTED', 'CHAIN_STEP_REJECTED'},
        'access_granted': {'ACCESS_REQUEST_APPROVED'},
        'access_denied': {'ACCESS_REQUEST_REJECTED'},
        'file_movements': {'FILE_SENT', 'DOCUMENT_FORWARDED', 'MOVEMENT_CLOSED'},
        'file_actions': {'FILE_CREATED', 'FILE_UPDATED', 'FILE_CLOSED', 'FILE_ARCHIVED', 'FILE_ACTIVATED'},
    }

    def get(self, request):
        import csv
        from datetime import timedelta
        from django.http import HttpResponse as HR

        staff = getattr(request.user, 'staff', None)
        is_registry = staff and staff.is_registry

        target_user = request.user
        user_id = request.GET.get('user_id')
        if user_id and is_registry:
            try:
                target_user = CustomUser.objects.get(pk=user_id)
            except CustomUser.DoesNotExist:
                pass

        days = int(request.GET.get('days', 30))
        since = timezone.now() - timedelta(days=days)

        qs = AuditLogEntry.objects.filter(timestamp__gte=since).select_related(
            'user', 'user__staff__department', 'user__staff__designation', 'content_type'
        ).order_by('-timestamp')

        # Filter by section if requested
        section = request.GET.get('section')
        if section in self.SECTION_MAP:
            qs = qs.filter(action__in=self.SECTION_MAP[section])

        # If not registry, or if a specific user was requested, filter by user
        if not is_registry or user_id:
            qs = qs.filter(user=target_user)

        fmt = request.GET.get('format', 'csv')

        rows = []
        for e in qs:
            d = e.details or {}
            u = e.user
            s = getattr(u, 'staff', None) if u else None
            
            # Identify the affected object
            obj_desc = ""
            if e.content_type:
                obj_desc = f"{e.content_type.model.title()} #{e.object_id}"
            
            # Special handling for common details
            file_info = d.get('file', d.get('file_number', ''))
            if d.get('file_title'):
                file_info = f"{file_info} ({d.get('file_title')})" if file_info else d.get('file_title')
            
            rows.append({
                'Date/Time': e.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                'Action': e.get_action_display(),
                'User': u.get_full_name() if u else 'System',
                'Username': u.username if u else '',
                'Staff ID': s.pk if s else '',
                'Department': s.department.name if s and s.department else '',
                'Designation': s.designation.name if s and s.designation else '',
                'Affected Object': obj_desc,
                'File Info': file_info,
                'Related User': d.get('requested_by', d.get('target_user', '')),
                'Reason/Details': d.get('denial_reason', d.get('reason', d.get('notes', ''))),
                'IP Address': e.ip_address or '',
                'Full Details (JSON)': ', '.join(f'{k}: {v}' for k, v in d.items()),
                'User Agent': e.user_agent or '',
            })

        # Construct a descriptive filename
        user_info = ""
        if target_user and (not is_registry or user_id):
            s = getattr(target_user, 'staff', None)
            name_part = (target_user.get_full_name() or target_user.username).replace(" ", "_")
            parts = [name_part]
            if s:
                if s.designation: parts.append(s.designation.name.replace(" ", "_"))
                if s.department: parts.append(s.department.name.replace(" ", "_"))
            user_info = "_" + "_".join(parts)

        filename = f"{section + '_' if section else ''}activity_log{user_info}_{timezone.now().strftime('%Y%m%d_%H%M')}"

        if fmt == 'excel':
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Activity Log'
                if rows:
                    headers = list(rows[0].keys())
                    ws.append(headers)
                    for r in rows:
                        ws.append(list(r.values()))
                    
                    # Basic styling
                    for cell in ws[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                
                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)
                response = HR(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                return response
            except ImportError:
                pass

        response = HR(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        if rows:
            writer = csv.DictWriter(response, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        return response
